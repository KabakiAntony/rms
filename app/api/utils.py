"""
it will hold the reusable
functions.
"""
import re
import os
import jwt
import random
import string
import datetime
from openpyxl import load_workbook
from pandas import read_excel
from sqlalchemy import create_engine
from functools import wraps
from flask import jsonify, make_response, abort, request
from app.api.model.company import Company, company_schema
from app.api.model.user import User, user_schema


KEY = os.getenv("SECRET_KEY")
DB_URL = os.environ.get("DATABASE_URL")
ALLOWED_EXTENSIONS = {"xlsx", "xls"}


def allowed_extension(filename):
    """check for allowed extensions"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def rename_file(filePath, company_id, upload_dir, file_type_str):
    """
    rename the file and save it
    add a company name and what the
    file is for and a now component
    in terms of date and time
    upload_dir is either employee_upload_folder
    or budget_upload_folder
    file_type is either _employees_ or _budget_
    """
    current_date = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
    company_ = Company.query.filter_by(id=company_id).first()
    this_company = company_schema.dump(company_)
    company_name = this_company["company"]
    new_file_path = (
        upload_dir + company_name + file_type_str + str(current_date) + ".xlsx"
    )
    os.rename(filePath, new_file_path)
    return new_file_path


def add_id_and_company_id(filePath, company_id):
    """
    add a user id and company id
    to the file then save it
    """
    workBook = load_workbook(filePath)
    sheet = workBook.active
    rows = sheet.max_row
    for row in range(2, rows + 1):
        sheet.cell(row, 1).value = generate_db_ids()
        sheet.cell(row, 2).value = company_id
    workBook.save(filePath)
    return filePath


def convert_to_csv(filePath, upload_dir):
    """convert excel file to csv"""
    # read the excel file
    dataFile = read_excel(filePath, engine="openpyxl")
    base_name = os.path.basename(filePath)
    csv_file_name = os.path.splitext(base_name)[0]
    csv_file_path = upload_dir + csv_file_name + ".csv"
    # convert to csv
    dataFile.to_csv(csv_file_path, index=False)
    return csv_file_path


def insert_csv(csv_file, db_relation):
    """
    read the csv file and insert
    contents to database db_relation
    is the database table we are
    saving data to it takes the form
    'public."Employees"'
    """
    engine = create_engine(DB_URL)
    konnection = engine.raw_connection()
    kursor = konnection.cursor()
    with open(csv_file, "r") as f:
        next(f)
        kursor.copy_from(f, db_relation, sep=",")
    konnection.commit()


def custom_make_response(key, message, status):
    """
    this is a custom make response for readability
    to avoid repeating the parts that make up a return
    message.
    """
    raw_dict = {"status": status}
    raw_dict[key] = message
    return make_response(jsonify(raw_dict), status)


def check_model_return(returned):
    """
    checks whether there was a return from the models
    assigns a 404/200 accordingly or 201 in the case of creation
    and assigns it a better message for the user
    """
    if not returned:
        message = "the resource you are looking for was not found!"
        status = 404
        response = custom_make_response("error", message, status)
    else:
        message = returned
        status = 200
        response = custom_make_response("data", message, status)
    return response


def isValidPassword(password):
    """
    check if the supplied password meets the
    expectations.
    """
    if len(password) < 6 or len(password) > 20:
        abort(
            custom_make_response(
                "error",
                "Password should be atleast 6 characters & not more than 20",
                400,
            )
        )

    lowercase_reg = re.search("[a-z]", password)
    uppercase_reg = re.search("[A-Z]", password)
    number_reg = re.search("[0-9]", password)

    if not lowercase_reg or not uppercase_reg or not number_reg:
        abort(
            custom_make_response(
                "error",
                "Password should contain at least 1 number,\
                 1 small letter & 1 Capital letter",
                400,
            )
        )


def check_for_whitespace(data, items_to_check):
    """
    check if the data supplied has whitespace
    """
    for key, value in data.items():
        if key in items_to_check and not value.strip():
            # exceptions go to site administrator log and email
            # the user gets a friendly error notification
            abort(custom_make_response("error", f"{key} cannot be left blank!", 400))
    return True


def token_required(f):
    """
    this token is used to allow the user to access
    certain routes
    """

    @wraps(f)
    def decorated(*args, **kwargs):
        user_token = None
        company_token = None
        if (
            request.cookies.get("auth_token")
            or request.args.get("in")
            or request.args.get("u")
        ):
            user_token = request.args.get("u") or request.cookies.get("auth_token")
            company_token = request.args.get("in")
        if not (user_token or company_token):
            return custom_make_response("error", "Token is missing", 401)
        try:
            if user_token:
                data = jwt.decode(user_token, KEY, algorithm="HS256")
                current_user = User.query.filter_by(id=data["id"]).first()
                _data = user_schema.dump(current_user)
            if company_token:
                data = jwt.decode(company_token, KEY, algorithms="HS256")
                current_company = Company.query.filter_by(
                    company=data["company"]
                ).first()
                _data = company_schema.dump(current_company)
        except Exception as e:
            # exceptions go to site administrator log and email
            # the user gets a friendly error notification
            return custom_make_response("error", f"Token {e}", 401)
        return f(_data, *args, **kwargs)

    return decorated


def generate_random_password():
    """
    generate random password for users
    signed up by admin, it is not used
    anywhere by for maintaining data
    sanity in the db
    """
    random_source = string.ascii_letters + string.digits + string.punctuation
    password = random.choice(string.ascii_lowercase)
    password += random.choice(string.ascii_uppercase)
    password += random.choice(string.digits)
    password += random.choice(string.punctuation)

    for i in range(6):
        password += random.choice(random_source)

    password_list = list(password)
    random.SystemRandom().shuffle(password_list)
    password = "".join(password_list)
    return password


def generate_db_ids():
    """
    we will use this function to generate unikue
    id for the database primary keys
    """
    unikueId = string.ascii_letters + string.digits
    unikueId = "".join(random.choice(unikueId) for i in range(10))

    return unikueId
