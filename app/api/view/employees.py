import os
import jwt
import datetime
from openpyxl import load_workbook
from pandas import read_excel
from app.api import rms
from app.api.model import db
from werkzeug.utils import secure_filename
from app.api.model.employees import Employees, employee_schema,\
    employees_schema
from app.api.model.company import Company, company_schema
from flask import request, abort
from app.api.utils import check_for_whitespace, custom_make_response,\
    generate_db_ids, token_required


KEY = os.environ.get('SECRET_KEY')
EMPLOYEE_UPLOAD_FOLDER = os.environ.get('EMPLOYEES_FOLDER')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_extension(filename):
    """check for allowed extensions"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def rename_file(filePath, company_id):
    """
    rename the file and save it
    add a company name and what the
    file is for and a now component
    in terms of date and time
    """
    current_date = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
    company_ = Company.query.filter_by(id=company_id).first()
    this_company = company_schema.dump(company_)
    company_name = this_company['company']
    new_file_path = EMPLOYEE_UPLOAD_FOLDER + company_name +\
        '_employees_' + str(current_date) + '.xlsx'
    os.rename(filePath, new_file_path)
    return new_file_path


def add_id_and_company_id(filePath, company_id):
    """
    add a user id and company id
    to the file then save it
    """
    workBook = load_workbook(filePath)
    sheet = workBook.active
    rows = sheet.max_row
    for row in range(2, rows+1):
        sheet.cell(row, 1).value = generate_db_ids()
        sheet.cell(row, 2).value = company_id
    workBook.save(filePath)
    return filePath


def to_csv_and_insert(filePath):
    """
    convert the to csv then
    save it to database
    """
    # konnection, kursor = db()
    dataFile = read_excel(filePath)
    base_name = os.path.basename(filePath)
    csv_file_name = os.path.splitext(base_name)[0]
    csv_file_path = EMPLOYEE_UPLOAD_FOLDER + csv_file_name + ".csv"
    dataFile.to_csv(csv_file_path, index=False)
    # with open(csv_file_path, 'r') as f:
    #     next(f)
    #     kursor.copy_from(f, 'Employees', sep=',')
    # db.session.commit()


@rms.route('/auth/upload/employees', methods=['POST'])
@token_required
def upload_employee_master(user):
    """
    upload the employee master file for a
    given company only the admin can do this
    """
    try:
        receivedFile = request.files['employeeExcelFile']
        if allowed_extension(receivedFile.filename) and receivedFile:
            secureFilename = secure_filename(receivedFile.filename)
            filePath = os.path.join(EMPLOYEE_UPLOAD_FOLDER, secureFilename)
            receivedFile.save(filePath)
            new_file_path = rename_file(filePath, user['companyId'])
            add_id_and_company_id(new_file_path, user['companyId'])
            updated_file_path = add_id_and_company_id(
                new_file_path, user['companyId']
            )
            to_csv_and_insert(updated_file_path)
            return custom_make_response(
                "data",
                "File uploaded successfully ",
                200
                )
        else:
            return custom_make_response(
                "error",
                "Only excel files are allowed !",
                400
            )
    except Exception as e:
        abort(
            custom_make_response(
                "error",
                f"{e}",
                # "An error occured uploading file \
                #     the administrator has been notified.",
                400
            )
        )
